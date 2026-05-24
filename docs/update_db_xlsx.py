"""
Update docs/db-design_document.xlsx to match 08-db-design.html (Ver 2.0)
- Fixes data types: DATETIME→TIMESTAMP, AUTO_INCREMENT→IDENTITY
- Adds all 24 tables
- Adds normalization evaluation sheet
- Updates complete index list
- Updates seed data
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX_PATH = "docs/db-design_document.xlsx"

# ── Style helpers ──────────────────────────────────────────────────────────────
BLUE_FILL   = PatternFill("solid", fgColor="1D4ED8")
TEAL_FILL   = PatternFill("solid", fgColor="0EA5A4")
GRAY_FILL   = PatternFill("solid", fgColor="F3F4F6")
GREEN_FILL  = PatternFill("solid", fgColor="D1FAE5")
AMBER_FILL  = PatternFill("solid", fgColor="FEF3C7")

def hdr(ws, row, col, value, fill=BLUE_FILL):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def sub_hdr(ws, row, col, value, fill=TEAL_FILL, cols=1):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+cols-1)
    return c

def cell(ws, row, col, value, bold=False, fill=None, center=False):
    c = ws.cell(row=row, column=col, value=value)
    if bold:
        c.font = Font(bold=True)
    if fill:
        c.fill = fill
    if center:
        c.alignment = Alignment(horizontal="center")
    return c

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_borders(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(row=r, column=c).border = thin_border()

# ── Load workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX_PATH)

# ── Rename sheets to proper Japanese ──────────────────────────────────────────
sheet_map = {0: "ER図", 1: "テーブル定義", 2: "インデックス一覧", 3: "初期データ"}
for idx, name in sheet_map.items():
    wb.worksheets[idx].title = name

# ── Add 正規化評価 sheet (insert at position 1) ────────────────────────────────
if "正規化評価" not in wb.sheetnames:
    ws_norm = wb.create_sheet("正規化評価", 1)
else:
    ws_norm = wb["正規化評価"]
    ws_norm.delete_rows(1, ws_norm.max_row)

# Header
ws_norm.merge_cells("A1:G1")
t = ws_norm["A1"]
t.value = "Pet Life Plus — データベース正規化評価"
t.font = Font(bold=True, size=14)
t.alignment = Alignment(horizontal="center")

# Section 1: 正規化チェック表
ws_norm.merge_cells("A3:G3")
sub_hdr(ws_norm, 3, 1, "正規化チェック表（全24テーブル）", TEAL_FILL, 7)

headers_norm = ["テーブル名", "1NF", "2NF", "3NF", "BCNF", "備考"]
for i, h in enumerate(headers_norm, 1):
    hdr(ws_norm, 4, i, h)

norm_data = [
    ("roles",                      "✅", "✅", "✅", "✅", "参照マスタ。依存関係なし"),
    ("users",                      "✅", "✅", "✅", "✅", "role_id FK のみ"),
    ("pets",                       "✅", "✅", "✅", "✅", "owner_user_id FK のみ"),
    ("health_records",             "✅", "✅", "✅", "✅", "pet_id FK のみ"),
    ("pet_care_records",           "✅", "✅", "✅", "✅", "pet_id FK のみ"),
    ("symptom_checks",             "✅", "✅", "✅", "✅", "pet_id FK のみ"),
    ("plans",                      "✅", "✅", "✅", "✅", "features_json は表示用（意図的冗長）"),
    ("plan_features",              "✅", "✅", "✅", "✅", "複合PK (plan_id, feature_code)"),
    ("subscriptions",              "✅", "✅", "⚠️", "✅", "user_id は pet_id 経由で導出可。性能目的で保持（意図的）"),
    ("invoices",                   "✅", "✅", "✅", "✅", "subscription_id FK のみ"),
    ("payments",                   "✅", "✅", "✅", "✅", "invoice_id FK のみ"),
    ("appointment_slots",          "✅", "✅", "✅", "✅", "参照マスタ"),
    ("appointments",               "✅", "✅", "✅", "✅", "user_id/pet_id/slot_id FK"),
    ("medical_histories",          "✅", "✅", "✅", "✅", "appointment_id UNIQUE（1:1）"),
    ("medical_attachments",        "✅", "✅", "✅", "✅", "medical_history_id FK"),
    ("consult_chat_messages",      "✅", "✅", "✅", "✅", "user_id FK のみ"),
    ("notifications",              "✅", "✅", "✅", "✅", "参照ヘッダ"),
    ("notification_recipients",    "✅", "✅", "✅", "✅", "複合PK (notification_id, user_id)"),
    ("email_templates",            "✅", "✅", "✅", "✅", "スタンドアロン"),
    ("email_messages",             "✅", "✅", "✅", "✅", "template_id FK"),
    ("pet_calendar_marks",         "✅", "✅", "✅", "✅", "pet_id FK"),
    ("dismissed_reminders",        "✅", "✅", "✅", "✅", "UNIQUE(user_id, reminder_key)"),
    ("announcements",              "✅", "✅", "✅", "✅", "スタンドアロン"),
    ("password_reset_tokens",      "✅", "✅", "✅", "✅", "user_id FK"),
]

for r, row in enumerate(norm_data, 5):
    for c, val in enumerate(row, 1):
        ws_norm.cell(row=r, column=c, value=val)

apply_borders(ws_norm, 4, 4+len(norm_data), 1, 6)

# Section 2: 意図的非正規化
row_offset = 5 + len(norm_data) + 1
ws_norm.merge_cells(f"A{row_offset}:F{row_offset}")
sub_hdr(ws_norm, row_offset, 1, "意図的非正規化の一覧", TEAL_FILL, 6)

headers_int = ["テーブル", "カラム", "非正規化の種類", "理由"]
for i, h in enumerate(headers_int, 1):
    hdr(ws_norm, row_offset+1, i, h)

intentional = [
    ("plans", "features_json",
     "3NF 違反（plan_features との重複）",
     "フロントエンド表示・マーケティング用途。plan_features は機能ゲート制御用"),
    ("subscriptions", "user_id",
     "3NF 軽度（pet_id→pets.owner_user_id で導出可）",
     "PlanFeatureMapper が user_id で直接クエリ。JOIN 排除で性能最適化"),
    ("appointments", "user_id",
     "pet_id→pets.owner_user_id で導出可",
     "予約検索・通知で頻繁にアクセス。JOIN コスト回避"),
]

for r, row in enumerate(intentional, row_offset+2):
    for c, val in enumerate(row, 1):
        ws_norm.cell(row=r, column=c, value=val)

apply_borders(ws_norm, row_offset+1, row_offset+1+len(intentional), 1, 4)

# Column widths
ws_norm.column_dimensions["A"].width = 28
ws_norm.column_dimensions["B"].width = 8
ws_norm.column_dimensions["C"].width = 8
ws_norm.column_dimensions["D"].width = 8
ws_norm.column_dimensions["E"].width = 8
ws_norm.column_dimensions["F"].width = 50

print("✅ 正規化評価 sheet done")

# ── テーブル定義 sheet ─────────────────────────────────────────────────────────
ws_tbl = wb["テーブル定義"]
ws_tbl.delete_rows(1, ws_tbl.max_row)

ws_tbl.merge_cells("A1:G1")
t = ws_tbl["A1"]
t.value = "Pet Life Plus — テーブル定義（全24テーブル）"
t.font = Font(bold=True, size=14)
t.alignment = Alignment(horizontal="center")

col_headers = ["テーブル名", "カラム名", "データ型", "NULL許可", "デフォルト値", "制約", "説明"]
for i, h in enumerate(col_headers, 1):
    hdr(ws_tbl, 3, i, h)

# All 24 tables definitions
# Format: (table, column, type, null, default, constraint, description)
tables = [
    # roles
    ("roles", "id",         "BIGINT",       "NO",  "IDENTITY",          "PK",               "ロールID"),
    ("roles", "role_code",  "VARCHAR(50)",  "NO",  "",                  "UNIQUE",            "ロールコード"),
    ("roles", "role_name",  "VARCHAR(100)", "NO",  "",                  "",                  "ロール名"),
    ("roles", "created_at", "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                  "作成日時"),
    ("roles", "updated_at", "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                  "更新日時"),
    # users
    ("users", "id",              "BIGINT",       "NO",  "IDENTITY",          "PK",               "ユーザーID"),
    ("users", "role_id",         "BIGINT",       "NO",  "",                  "FK → roles.id",    "ロールID"),
    ("users", "name",            "VARCHAR(100)", "NO",  "",                  "",                  "氏名"),
    ("users", "email",           "VARCHAR(255)", "NO",  "",                  "UNIQUE",            "メール"),
    ("users", "password_hash",   "VARCHAR(255)", "NO",  "",                  "",                  "パスワードハッシュ"),
    ("users", "phone",           "VARCHAR(20)",  "YES", "NULL",              "",                  "電話番号"),
    ("users", "status",          "VARCHAR(20)",  "NO",  "'ACTIVE'",          "",                  "状態"),
    ("users", "slack_user_id",   "VARCHAR(100)", "YES", "NULL",              "",                  "Slack ユーザーID"),
    ("users", "line_user_id",    "VARCHAR(100)", "YES", "NULL",              "",                  "LINE ユーザーID"),
    ("users", "last_login_at",   "TIMESTAMP",    "YES", "NULL",              "",                  "最終ログイン日時"),
    ("users", "deleted_at",      "TIMESTAMP",    "YES", "NULL",              "",                  "論理削除日時"),
    ("users", "created_at",      "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                  "作成日時"),
    ("users", "updated_at",      "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                  "更新日時"),
    # pets
    ("pets", "id",                  "BIGINT",       "NO",  "IDENTITY",          "PK",               "ペットID"),
    ("pets", "owner_user_id",       "BIGINT",       "NO",  "",                  "FK → users.id",    "オーナーID"),
    ("pets", "name",                "VARCHAR(100)", "NO",  "",                  "",                  "ペット名"),
    ("pets", "species",             "VARCHAR(30)",  "NO",  "",                  "",                  "種別"),
    ("pets", "breed",               "VARCHAR(100)", "YES", "NULL",              "",                  "品種"),
    ("pets", "sex",                 "VARCHAR(10)",  "YES", "NULL",              "",                  "性別"),
    ("pets", "birth_date",          "DATE",         "YES", "NULL",              "",                  "誕生日"),
    ("pets", "weight_baseline_kg",  "DECIMAL(5,2)", "YES", "NULL",              "",                  "基準体重"),
    ("pets", "image_path",          "VARCHAR(500)", "YES", "NULL",              "",                  "プロフィール画像パス"),
    ("pets", "deceased_at",         "TIMESTAMP",    "YES", "NULL",              "",                  "死亡日時"),
    ("pets", "deleted_at",          "TIMESTAMP",    "YES", "NULL",              "",                  "論理削除日時"),
    ("pets", "created_at",          "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                  "作成日時"),
    ("pets", "updated_at",          "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                  "更新日時"),
    # health_records
    ("health_records", "id",              "BIGINT",       "NO",  "IDENTITY",          "PK",              "記録ID"),
    ("health_records", "pet_id",          "BIGINT",       "NO",  "",                  "FK → pets.id",    "ペットID"),
    ("health_records", "record_date",     "DATE",         "NO",  "",                  "",                "記録日"),
    ("health_records", "weight_kg",       "DECIMAL(5,2)", "YES", "NULL",              "",                "体重"),
    ("health_records", "temperature",     "DECIMAL(4,1)", "YES", "NULL",              "",                "体温"),
    ("health_records", "memo",            "TEXT",         "YES", "NULL",              "",                "メモ"),
    ("health_records", "vaccine_name",    "VARCHAR(200)", "YES", "NULL",              "",                "ワクチン名"),
    ("health_records", "vaccine_date",    "DATE",         "YES", "NULL",              "",                "接種日"),
    ("health_records", "next_vaccine_at", "DATE",         "YES", "NULL",              "",                "次回接種日"),
    ("health_records", "deleted_at",      "TIMESTAMP",    "YES", "NULL",              "",                "論理削除日時"),
    ("health_records", "created_at",      "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
    ("health_records", "updated_at",      "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "更新日時"),
    # pet_care_records
    ("pet_care_records", "id",          "BIGINT",       "NO",  "IDENTITY",          "PK",              "記録ID"),
    ("pet_care_records", "pet_id",      "BIGINT",       "NO",  "",                  "FK → pets.id",    "ペットID"),
    ("pet_care_records", "care_type",   "VARCHAR(50)",  "NO",  "",                  "",                "ケア種別"),
    ("pet_care_records", "care_date",   "DATE",         "NO",  "",                  "",                "実施日"),
    ("pet_care_records", "memo",        "TEXT",         "YES", "NULL",              "",                "メモ"),
    ("pet_care_records", "deleted_at",  "TIMESTAMP",    "YES", "NULL",              "",                "論理削除日時"),
    ("pet_care_records", "created_at",  "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
    ("pet_care_records", "updated_at",  "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "更新日時"),
    # symptom_checks
    ("symptom_checks", "id",           "BIGINT",       "NO",  "IDENTITY",          "PK",              "チェックID"),
    ("symptom_checks", "pet_id",       "BIGINT",       "NO",  "",                  "FK → pets.id",    "ペットID"),
    ("symptom_checks", "user_id",      "BIGINT",       "NO",  "",                  "FK → users.id",   "ユーザーID"),
    ("symptom_checks", "symptoms",     "TEXT",         "NO",  "",                  "",                "症状テキスト"),
    ("symptom_checks", "ai_response",  "TEXT",         "YES", "NULL",              "",                "AI 回答"),
    ("symptom_checks", "severity",     "VARCHAR(20)",  "YES", "NULL",              "",                "重症度"),
    ("symptom_checks", "created_at",   "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
    # plans
    ("plans", "id",            "BIGINT",       "NO",  "IDENTITY",          "PK",              "プランID"),
    ("plans", "plan_code",     "VARCHAR(50)",  "NO",  "",                  "UNIQUE",           "プランコード"),
    ("plans", "plan_name",     "VARCHAR(100)", "NO",  "",                  "",                "プラン名"),
    ("plans", "price",         "INTEGER",      "NO",  "0",                 "",                "月額料金（円）"),
    ("plans", "features_json", "TEXT",         "YES", "NULL",              "",                "機能一覧JSON（表示用）"),
    ("plans", "deleted_at",    "TIMESTAMP",    "YES", "NULL",              "",                "論理削除日時"),
    ("plans", "created_at",    "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
    ("plans", "updated_at",    "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "更新日時"),
    # plan_features
    ("plan_features", "plan_id",       "BIGINT",      "NO",  "",  "PK, FK → plans.id",  "プランID"),
    ("plan_features", "feature_code",  "VARCHAR(50)", "NO",  "",  "PK",                  "機能コード"),
    # subscriptions
    ("subscriptions", "id",             "BIGINT",    "NO",  "IDENTITY",          "PK",                  "サブスクID"),
    ("subscriptions", "user_id",        "BIGINT",    "NO",  "",                  "FK → users.id",       "ユーザーID"),
    ("subscriptions", "pet_id",         "BIGINT",    "NO",  "",                  "FK → pets.id",        "ペットID"),
    ("subscriptions", "plan_id",        "BIGINT",    "NO",  "",                  "FK → plans.id",       "プランID"),
    ("subscriptions", "status",         "VARCHAR(20)","NO",  "'ACTIVE'",         "",                    "状態"),
    ("subscriptions", "start_date",     "DATE",      "NO",  "",                  "",                    "開始日"),
    ("subscriptions", "end_date",       "DATE",      "YES", "NULL",              "",                    "終了日"),
    ("subscriptions", "renewal_date",   "DATE",      "YES", "NULL",              "",                    "更新日"),
    ("subscriptions", "deleted_at",     "TIMESTAMP", "YES", "NULL",              "",                    "論理削除日時"),
    ("subscriptions", "created_at",     "TIMESTAMP", "NO",  "CURRENT_TIMESTAMP", "",                    "作成日時"),
    ("subscriptions", "updated_at",     "TIMESTAMP", "NO",  "CURRENT_TIMESTAMP", "",                    "更新日時"),
    # invoices
    ("invoices", "id",              "BIGINT",       "NO",  "IDENTITY",          "PK",                        "請求書ID"),
    ("invoices", "subscription_id", "BIGINT",       "NO",  "",                  "FK → subscriptions.id",     "サブスクID"),
    ("invoices", "amount",          "INTEGER",      "NO",  "",                  "",                          "金額（円）"),
    ("invoices", "status",          "VARCHAR(20)",  "NO",  "'UNPAID'",          "",                          "状態"),
    ("invoices", "due_date",        "DATE",         "NO",  "",                  "",                          "支払期限"),
    ("invoices", "paid_at",         "TIMESTAMP",    "YES", "NULL",              "",                          "支払日時"),
    ("invoices", "deleted_at",      "TIMESTAMP",    "YES", "NULL",              "",                          "論理削除日時"),
    ("invoices", "created_at",      "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                          "作成日時"),
    ("invoices", "updated_at",      "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                          "更新日時"),
    # payments
    ("payments", "id",             "BIGINT",       "NO",  "IDENTITY",          "PK",                   "支払ID"),
    ("payments", "invoice_id",     "BIGINT",       "NO",  "",                  "FK → invoices.id",     "請求書ID"),
    ("payments", "amount",         "INTEGER",      "NO",  "",                  "",                     "支払金額（円）"),
    ("payments", "payment_method", "VARCHAR(50)",  "YES", "NULL",              "",                     "支払方法"),
    ("payments", "paid_at",        "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                     "支払日時"),
    ("payments", "created_at",     "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                     "作成日時"),
    # appointment_slots
    ("appointment_slots", "id",          "BIGINT",    "NO",  "IDENTITY",          "PK",              "枠ID"),
    ("appointment_slots", "slot_date",   "DATE",      "NO",  "",                  "",                "予約日"),
    ("appointment_slots", "start_time",  "TIME",      "NO",  "",                  "",                "開始時刻"),
    ("appointment_slots", "end_time",    "TIME",      "NO",  "",                  "",                "終了時刻"),
    ("appointment_slots", "max_count",   "INTEGER",   "NO",  "1",                 "",                "最大受付数"),
    ("appointment_slots", "deleted_at",  "TIMESTAMP", "YES", "NULL",              "",                "論理削除日時"),
    ("appointment_slots", "created_at",  "TIMESTAMP", "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
    ("appointment_slots", "updated_at",  "TIMESTAMP", "NO",  "CURRENT_TIMESTAMP", "",                "更新日時"),
    # appointments
    ("appointments", "id",             "BIGINT",       "NO",  "IDENTITY",          "PK",                         "予約ID"),
    ("appointments", "user_id",        "BIGINT",       "NO",  "",                  "FK → users.id",              "ユーザーID"),
    ("appointments", "pet_id",         "BIGINT",       "NO",  "",                  "FK → pets.id",               "ペットID"),
    ("appointments", "slot_id",        "BIGINT",       "YES", "NULL",              "FK → appointment_slots.id",  "枠ID"),
    ("appointments", "appointment_date","DATE",         "NO",  "",                  "",                           "予約日"),
    ("appointments", "appointment_time","TIME",         "NO",  "",                  "",                           "予約時刻"),
    ("appointments", "reason",         "TEXT",         "YES", "NULL",              "",                           "受診理由"),
    ("appointments", "status",         "VARCHAR(20)",  "NO",  "'PENDING'",         "",                           "状態"),
    ("appointments", "zoom_join_url",  "VARCHAR(500)", "YES", "NULL",              "",                           "Zoom 参加URL"),
    ("appointments", "deleted_at",     "TIMESTAMP",    "YES", "NULL",              "",                           "論理削除日時"),
    ("appointments", "created_at",     "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                           "作成日時"),
    ("appointments", "updated_at",     "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                           "更新日時"),
    # medical_histories
    ("medical_histories", "id",             "BIGINT",    "NO",  "IDENTITY",          "PK",                    "診療記録ID"),
    ("medical_histories", "appointment_id", "BIGINT",    "YES", "NULL",              "FK → appointments.id UNIQUE","予約ID"),
    ("medical_histories", "pet_id",         "BIGINT",    "NO",  "",                  "FK → pets.id",          "ペットID"),
    ("medical_histories", "user_id",        "BIGINT",    "NO",  "",                  "FK → users.id",         "担当獣医ID"),
    ("medical_histories", "visit_date",     "DATE",      "NO",  "",                  "",                      "来院日"),
    ("medical_histories", "diagnosis",      "TEXT",      "YES", "NULL",              "",                      "診断"),
    ("medical_histories", "treatment",      "TEXT",      "YES", "NULL",              "",                      "処置"),
    ("medical_histories", "prescription",   "TEXT",      "YES", "NULL",              "",                      "処方"),
    ("medical_histories", "notes",          "TEXT",      "YES", "NULL",              "",                      "備考"),
    ("medical_histories", "deleted_at",     "TIMESTAMP", "YES", "NULL",              "",                      "論理削除日時"),
    ("medical_histories", "created_at",     "TIMESTAMP", "NO",  "CURRENT_TIMESTAMP", "",                      "作成日時"),
    ("medical_histories", "updated_at",     "TIMESTAMP", "NO",  "CURRENT_TIMESTAMP", "",                      "更新日時"),
    # medical_attachments
    ("medical_attachments", "id",                 "BIGINT",       "NO",  "IDENTITY",          "PK",                         "添付ID"),
    ("medical_attachments", "medical_history_id", "BIGINT",       "NO",  "",                  "FK → medical_histories.id",  "診療記録ID"),
    ("medical_attachments", "file_path",          "VARCHAR(500)", "NO",  "",                  "",                          "ファイルパス"),
    ("medical_attachments", "file_type",          "VARCHAR(50)",  "YES", "NULL",              "",                          "ファイル種別"),
    ("medical_attachments", "created_at",         "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                          "作成日時"),
    # consult_chat_messages
    ("consult_chat_messages", "id",          "BIGINT",      "NO",  "IDENTITY",          "PK",              "メッセージID"),
    ("consult_chat_messages", "user_id",     "BIGINT",      "NO",  "",                  "FK → users.id",   "ユーザーID"),
    ("consult_chat_messages", "channel",     "VARCHAR(20)", "NO",  "",                  "",                "チャンネル種別"),
    ("consult_chat_messages", "sender_type", "VARCHAR(20)", "NO",  "",                  "",                "送信者種別"),
    ("consult_chat_messages", "content",     "TEXT",        "NO",  "",                  "",                "メッセージ本文"),
    ("consult_chat_messages", "created_at",  "TIMESTAMP",   "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
    # notifications
    ("notifications", "id",          "BIGINT",       "NO",  "IDENTITY",          "PK",              "通知ID"),
    ("notifications", "type",        "VARCHAR(50)",  "NO",  "",                  "",                "通知種別"),
    ("notifications", "title",       "VARCHAR(200)", "NO",  "",                  "",                "タイトル"),
    ("notifications", "body",        "TEXT",         "YES", "NULL",              "",                "本文"),
    ("notifications", "created_at",  "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
    # notification_recipients
    ("notification_recipients", "notification_id", "BIGINT",    "NO",  "",                  "PK, FK → notifications.id",  "通知ID"),
    ("notification_recipients", "user_id",         "BIGINT",    "NO",  "",                  "PK, FK → users.id",          "ユーザーID"),
    ("notification_recipients", "is_read",         "BOOLEAN",   "NO",  "FALSE",             "",                           "既読フラグ"),
    ("notification_recipients", "read_at",         "TIMESTAMP", "YES", "NULL",              "",                           "既読日時"),
    # email_templates
    ("email_templates", "id",           "BIGINT",       "NO",  "IDENTITY",          "PK",              "テンプレートID"),
    ("email_templates", "template_code","VARCHAR(50)",  "NO",  "",                  "UNIQUE",           "テンプレートコード"),
    ("email_templates", "subject",      "VARCHAR(255)", "NO",  "",                  "",                "件名"),
    ("email_templates", "body",         "TEXT",         "NO",  "",                  "",                "本文テンプレート"),
    ("email_templates", "created_at",   "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
    ("email_templates", "updated_at",   "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "更新日時"),
    # email_messages
    ("email_messages", "id",          "BIGINT",       "NO",  "IDENTITY",          "PK",                       "メールID"),
    ("email_messages", "template_id", "BIGINT",       "YES", "NULL",              "FK → email_templates.id",  "テンプレートID"),
    ("email_messages", "to_address",  "VARCHAR(255)", "NO",  "",                  "",                         "宛先"),
    ("email_messages", "subject",     "VARCHAR(255)", "NO",  "",                  "",                         "件名"),
    ("email_messages", "body",        "TEXT",         "NO",  "",                  "",                         "本文"),
    ("email_messages", "status",      "VARCHAR(20)",  "NO",  "'PENDING'",         "",                         "送信状態"),
    ("email_messages", "sent_at",     "TIMESTAMP",    "YES", "NULL",              "",                         "送信日時"),
    ("email_messages", "created_at",  "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                         "作成日時"),
    # pet_calendar_marks
    ("pet_calendar_marks", "id",         "BIGINT",       "NO",  "IDENTITY",          "PK",              "マークID"),
    ("pet_calendar_marks", "pet_id",     "BIGINT",       "NO",  "",                  "FK → pets.id",    "ペットID"),
    ("pet_calendar_marks", "mark_date",  "DATE",         "NO",  "",                  "",                "マーク日"),
    ("pet_calendar_marks", "label",      "VARCHAR(100)", "YES", "NULL",              "",                "ラベル"),
    ("pet_calendar_marks", "color",      "VARCHAR(20)",  "YES", "NULL",              "",                "色"),
    ("pet_calendar_marks", "deleted_at", "TIMESTAMP",    "YES", "NULL",              "",                "論理削除日時"),
    ("pet_calendar_marks", "created_at", "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
    ("pet_calendar_marks", "updated_at", "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "更新日時"),
    # dismissed_reminders
    ("dismissed_reminders", "id",           "BIGINT",       "NO",  "IDENTITY",          "PK",              "ID"),
    ("dismissed_reminders", "user_id",      "BIGINT",       "NO",  "",                  "FK → users.id",   "ユーザーID"),
    ("dismissed_reminders", "reminder_key", "VARCHAR(200)", "NO",  "",                  "UNIQUE(user_id,reminder_key)", "リマインダーキー"),
    ("dismissed_reminders", "created_at",   "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
    # announcements
    ("announcements", "id",           "BIGINT",       "NO",  "IDENTITY",          "PK",              "お知らせID"),
    ("announcements", "title",        "VARCHAR(200)", "NO",  "",                  "",                "タイトル"),
    ("announcements", "body",         "TEXT",         "NO",  "",                  "",                "本文"),
    ("announcements", "published_at", "TIMESTAMP",    "YES", "NULL",              "",                "公開日時"),
    ("announcements", "deleted_at",   "TIMESTAMP",    "YES", "NULL",              "",                "論理削除日時"),
    ("announcements", "created_at",   "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
    ("announcements", "updated_at",   "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "更新日時"),
    # password_reset_tokens
    ("password_reset_tokens", "id",          "BIGINT",       "NO",  "IDENTITY",          "PK",              "ID"),
    ("password_reset_tokens", "user_id",     "BIGINT",       "NO",  "",                  "FK → users.id",   "ユーザーID"),
    ("password_reset_tokens", "token",       "VARCHAR(255)", "NO",  "",                  "UNIQUE",           "トークン"),
    ("password_reset_tokens", "expires_at",  "TIMESTAMP",    "NO",  "",                  "",                "有効期限"),
    ("password_reset_tokens", "used_at",     "TIMESTAMP",    "YES", "NULL",              "",                "使用日時"),
    ("password_reset_tokens", "created_at",  "TIMESTAMP",    "NO",  "CURRENT_TIMESTAMP", "",                "作成日時"),
]

current_table = None
row_num = 4
for entry in tables:
    tbl_name = entry[0]
    if tbl_name != current_table:
        # Table group header
        current_table = tbl_name
        ws_tbl.merge_cells(f"A{row_num}:G{row_num}")
        c = ws_tbl.cell(row=row_num, column=1, value=f"▶ {tbl_name}")
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = TEAL_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        row_num += 1

    for col_idx, val in enumerate(entry, 1):
        ws_tbl.cell(row=row_num, column=col_idx, value=val)
    row_num += 1

apply_borders(ws_tbl, 3, row_num-1, 1, 7)

ws_tbl.column_dimensions["A"].width = 26
ws_tbl.column_dimensions["B"].width = 22
ws_tbl.column_dimensions["C"].width = 16
ws_tbl.column_dimensions["D"].width = 10
ws_tbl.column_dimensions["E"].width = 26
ws_tbl.column_dimensions["F"].width = 28
ws_tbl.column_dimensions["G"].width = 26

print(f"✅ テーブル定義 sheet done ({row_num-4} rows)")

# ── インデックス一覧 sheet ────────────────────────────────────────────────────
ws_idx = wb["インデックス一覧"]
ws_idx.delete_rows(1, ws_idx.max_row)

ws_idx.merge_cells("A1:E1")
t = ws_idx["A1"]
t.value = "Pet Life Plus — インデックス設計"
t.font = Font(bold=True, size=14)
t.alignment = Alignment(horizontal="center")

idx_headers = ["テーブル名", "インデックス名", "対象カラム", "種別", "設定理由"]
for i, h in enumerate(idx_headers, 1):
    hdr(ws_idx, 3, i, h)

# Group headers + index entries
# Format: (None, group_label) for group header, or (table, index, col, type, reason) for data
idx_entries = [
    (None, "Auth（認証）"),
    ("users", "uk_users_email",       "email",           "UNIQUE", "ログインID の一意性確保"),
    ("users", "idx_users_role_id",    "role_id",         "INDEX",  "ロール別検索・FK 結合"),
    ("users", "idx_users_slack",      "slack_user_id",   "INDEX",  "Slack イベント受信時のユーザー解決"),
    ("users", "idx_users_line",       "line_user_id",    "INDEX",  "LINE イベント受信時のユーザー解決"),

    (None, "ペット・健康管理"),
    ("pets", "idx_pets_owner",             "owner_user_id",    "INDEX",  "オーナー別ペット一覧検索"),
    ("health_records", "idx_hr_pet_date",  "pet_id, record_date", "INDEX", "ペット別健康記録の日付ソート"),
    ("pet_care_records", "idx_pcr_pet",    "pet_id",           "INDEX",  "ペット別ケア記録検索"),
    ("symptom_checks", "idx_sc_pet",       "pet_id",           "INDEX",  "ペット別症状チェック履歴"),
    ("symptom_checks", "idx_sc_user",      "user_id",          "INDEX",  "ユーザー別症状チェック履歴"),

    (None, "プラン・課金"),
    ("plans", "uk_plans_code",             "plan_code",        "UNIQUE", "プランコードの一意性確保"),
    ("subscriptions", "idx_sub_user",      "user_id",          "INDEX",  "ユーザー別サブスクリプション検索"),
    ("subscriptions", "idx_sub_pet",       "pet_id",           "INDEX",  "ペット別サブスクリプション検索"),
    ("subscriptions", "idx_sub_plan",      "plan_id",          "INDEX",  "プラン別サブスクリプション集計"),
    ("subscriptions", "idx_sub_renewal",   "renewal_date",     "INDEX",  "更新日リマインダー処理"),
    ("invoices", "idx_inv_subscription",   "subscription_id",  "INDEX",  "サブスクリプション別請求書検索"),
    ("payments", "idx_pay_invoice",        "invoice_id",       "INDEX",  "請求書別支払履歴検索"),

    (None, "予約・診療記録"),
    ("appointment_slots", "idx_slots_date", "slot_date",       "INDEX",  "日付別予約枠検索"),
    ("appointments", "idx_appt_user",       "user_id",         "INDEX",  "ユーザー別予約一覧"),
    ("appointments", "idx_appt_pet",        "pet_id",          "INDEX",  "ペット別予約履歴"),
    ("appointments", "idx_appt_slot",       "slot_id",         "INDEX",  "枠別予約数カウント"),
    ("appointments", "idx_appt_date",       "appointment_date","INDEX",  "日付別予約一覧（カレンダー表示）"),
    ("medical_histories", "idx_mh_pet",     "pet_id",          "INDEX",  "ペット別診療記録一覧"),
    ("medical_histories", "idx_mh_appt",    "appointment_id",  "UNIQUE", "予約との1:1整合性保証"),
    ("medical_attachments", "idx_ma_mh",    "medical_history_id", "INDEX", "診療記録別添付ファイル検索"),

    (None, "メッセージング・通知"),
    ("consult_chat_messages", "idx_ccm_user",    "user_id",          "INDEX",  "ユーザー別チャット履歴"),
    ("notifications", "idx_notif_type",          "type",             "INDEX",  "通知種別フィルター"),
    ("notification_recipients", "idx_nr_user",   "user_id",          "INDEX",  "ユーザー別未読通知検索"),
    ("notification_recipients", "idx_nr_read",   "user_id, is_read", "INDEX",  "未読フィルター最適化"),
    ("email_templates", "uk_et_code",            "template_code",    "UNIQUE", "テンプレートコードの一意性確保"),
    ("email_messages", "idx_em_status",          "status",           "INDEX",  "送信待ちキュー処理"),
    ("email_messages", "idx_em_template",        "template_id",      "INDEX",  "テンプレート別送信履歴"),

    (None, "UX補助"),
    ("pet_calendar_marks", "idx_pcm_pet_date",    "pet_id, mark_date","INDEX",  "ペット別カレンダーマーク検索"),
    ("dismissed_reminders", "uk_dr_user_key",     "user_id, reminder_key", "UNIQUE", "ユーザー別リマインダー重複防止"),
    ("announcements", "idx_ann_published",        "published_at",     "INDEX",  "公開日時順ソート"),
    ("password_reset_tokens", "uk_prt_token",     "token",            "UNIQUE", "トークンの一意性確保"),
    ("password_reset_tokens", "idx_prt_user",     "user_id",          "INDEX",  "ユーザー別トークン検索"),
]

row_num = 4
for entry in idx_entries:
    if entry[0] is None:
        # Group header
        ws_idx.merge_cells(f"A{row_num}:E{row_num}")
        c = ws_idx.cell(row=row_num, column=1, value=f"── {entry[1]} ──")
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = TEAL_FILL
        c.alignment = Alignment(horizontal="left")
    else:
        for col_idx, val in enumerate(entry, 1):
            ws_idx.cell(row=row_num, column=col_idx, value=val)
    row_num += 1

apply_borders(ws_idx, 3, row_num-1, 1, 5)

ws_idx.column_dimensions["A"].width = 26
ws_idx.column_dimensions["B"].width = 32
ws_idx.column_dimensions["C"].width = 28
ws_idx.column_dimensions["D"].width = 10
ws_idx.column_dimensions["E"].width = 40

print(f"✅ インデックス一覧 sheet done ({row_num-4} rows)")

# ── 初期データ sheet ──────────────────────────────────────────────────────────
ws_data = wb["初期データ"]
ws_data.delete_rows(1, ws_data.max_row)

ws_data.merge_cells("A1:D1")
t = ws_data["A1"]
t.value = "Pet Life Plus — 初期データ（data.sql）"
t.font = Font(bold=True, size=14)
t.alignment = Alignment(horizontal="center")

# Plans
ws_data.merge_cells("A3:D3")
sub_hdr(ws_data, 3, 1, "plans（プラン）", TEAL_FILL, 4)
for i, h in enumerate(["plan_code", "plan_name", "price（円/月）", "features_json（表示用）"], 1):
    hdr(ws_data, 4, i, h)
plan_rows = [
    ("LIGHT",    "ライトプラン",    "980",  "なし"),
    ("STANDARD", "スタンダードプラン","1980", "AI症状チェック、Slackbot、LINEbot"),
    ("PREMIUM",  "プレミアムプラン", "2980", "AI症状チェック、Slackbot、LINEbot、Zoomオンライン診療"),
]
for r, row in enumerate(plan_rows, 5):
    for c, v in enumerate(row, 1):
        ws_data.cell(row=r, column=c, value=v)

# plan_features
ws_data.merge_cells("A9:D9")
sub_hdr(ws_data, 9, 1, "plan_features（機能コード）", TEAL_FILL, 4)
for i, h in enumerate(["plan_code", "feature_code", "説明", ""], 1):
    hdr(ws_data, 10, i, h)
pf_rows = [
    ("STANDARD", "AI_SYMPTOM",   "AI症状チェック機能", ""),
    ("STANDARD", "SLACK_BOT",    "Slackbot相談機能",    ""),
    ("STANDARD", "LINE_BOT",     "LINEbot相談機能",     ""),
    ("PREMIUM",  "AI_SYMPTOM",   "AI症状チェック機能", ""),
    ("PREMIUM",  "SLACK_BOT",    "Slackbot相談機能",    ""),
    ("PREMIUM",  "LINE_BOT",     "LINEbot相談機能",     ""),
    ("PREMIUM",  "ZOOM_CONSULT", "Zoomオンライン診療",  ""),
]
for r, row in enumerate(pf_rows, 11):
    for c, v in enumerate(row, 1):
        ws_data.cell(row=r, column=c, value=v)

# Users (created by DataInitializer with BCrypt)
ws_data.merge_cells("A20:D20")
sub_hdr(ws_data, 20, 1, "users（DataInitializerが BCrypt ハッシュで作成）", TEAL_FILL, 4)
for i, h in enumerate(["email", "平文パスワード", "ロール", "備考"], 1):
    hdr(ws_data, 21, i, h)
user_rows = [
    ("admin@petlifeplus.local",    "admin123",    "ADMIN", "管理者"),
    ("owner1@petlifeplus.local",   "user123",     "USER",  "一般ユーザー（プレミアム）"),
    ("vet1@petlifeplus.local",     "vet123",      "VET",   "獣医師"),
    ("staff1@petlifeplus.local",   "staff123",    "STAFF", "スタッフ"),
    ("owner2@petlifeplus.local",   "user123",     "USER",  "一般ユーザー（スタンダード）"),
    ("owner.light@petlifeplus.local",    "light123",    "USER",  "ライトプランユーザー"),
    ("owner.standard@petlifeplus.local", "standard123", "USER",  "スタンダードプランユーザー"),
    ("owner.premium@petlifeplus.local",  "premium123",  "USER",  "プレミアムプランユーザー"),
]
for r, row in enumerate(user_rows, 22):
    for c, v in enumerate(row, 1):
        ws_data.cell(row=r, column=c, value=v)

# Other seed data summary
ws_data.merge_cells("A32:D32")
sub_hdr(ws_data, 32, 1, "その他の初期データ（data.sql）", TEAL_FILL, 4)
for i, h in enumerate(["テーブル", "件数", "内容", ""], 1):
    hdr(ws_data, 33, i, h)
other_rows = [
    ("pets",          "6",  "各ユーザーの飼いペット（犬・猫）", ""),
    ("health_records","3",  "ペット1〜3の健康記録",              ""),
    ("pet_care_records","3","ペット1〜3のケア記録",              ""),
    ("subscriptions", "5",  "各プランへの契約",                  ""),
]
for r, row in enumerate(other_rows, 34):
    for c, v in enumerate(row, 1):
        ws_data.cell(row=r, column=c, value=v)

apply_borders(ws_data, 4, 8, 1, 4)
apply_borders(ws_data, 10, 17, 1, 4)
apply_borders(ws_data, 21, 29, 1, 4)
apply_borders(ws_data, 33, 37, 1, 4)

ws_data.column_dimensions["A"].width = 36
ws_data.column_dimensions["B"].width = 18
ws_data.column_dimensions["C"].width = 22
ws_data.column_dimensions["D"].width = 30

print("✅ 初期データ sheet done")

# ── ER図 sheet — update title ─────────────────────────────────────────────────
ws_er = wb["ER図"]
ws_er["A1"] = "Pet Life Plus ER図（テキスト形式）"
ws_er["A1"].font = Font(bold=True, size=14)

# Update ver note
ws_er["A2"] = "Ver 2.0 — 2026-05-24 更新 | 全24テーブル"

print("✅ ER図 sheet updated")

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(XLSX_PATH)
print(f"\n✅ Saved: {XLSX_PATH}")
